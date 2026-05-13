"""
One-time script: adds Tahoe Live and SLC Live rows to the Outreach Tracker
and renames the 'Resort Name' column header to 'Resort or Venue'.

Run via the add-live-venues GitHub Actions workflow.
"""

import io
import os
import base64

import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

TRACKER_FILE_ID = os.environ["GDRIVE_TRACKER_FILE_ID"]
ARCHIVE_FOLDER_ID = os.environ["GDRIVE_ARCHIVE_FOLDER_ID"]
TRACKER_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _google_creds():
    creds = Credentials(
        token=None,
        refresh_token=os.environ["GOOGLE_REFRESH_TOKEN"],
        client_id=os.environ["GOOGLE_CLIENT_ID"],
        client_secret=os.environ["GOOGLE_CLIENT_SECRET"],
        token_uri="https://oauth2.googleapis.com/token",
        scopes=[
            "https://mail.google.com/",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    creds.refresh(Request())
    return creds


def main():
    creds = _google_creds()
    drive = build("drive", "v3", credentials=creds)

    # Download tracker
    print("Downloading tracker...")
    request = drive.files().get_media(fileId=TRACKER_FILE_ID)
    buf = io.BytesIO()
    from googleapiclient.http import MediaIoBaseDownload
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    xlsx_bytes = buf.getvalue()

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    sheet = wb["Outreach Tracker"]

    # 1. Rename header
    sheet.cell(row=1, column=1).value = "Resort or Venue"
    print("Renamed header to 'Resort or Venue'")

    # 2. Find last data row
    last_row = sheet.max_row
    while last_row > 1 and sheet.cell(row=last_row, column=1).value is None:
        last_row -= 1

    # Column map (1-based):
    # 1 Resort or Venue, 2 State, 3 Region, 4 Tier, 5 Books DJs/Music?,
    # 6 Season Open, 7 Season Close, 8 Contact Name, 9 Contact Email,
    # 10 Contact Phone, 11 Outreach Type, 12 Status,
    # 13 Date First Contacted, 14 Date Last Contact,
    # 15 Next Follow-up, 16 Preferred Tour Window, 17 Booked Date(s), 18 Notes

    new_rows = [
        {
            1:  "Tahoe Live (Palisades Tahoe)",
            2:  "CA",
            3:  "California",
            4:  1,
            5:  "Yes",
            8:  "Dustin",
            9:  "dustin@parkcitylive.us",
            11: "Cold",
            12: "Email Sent",
            13: "2026-05-13",
            14: "2026-05-13",
            15: "2026-05-20",
            18: "Festival/event organizer at Palisades Tahoe. Pitched side-stage programming between mainstage acts (snowcat stage as second discovery moment). Same contact manages SLC Live. Subject: 'Disco Carrot @ Tahoe Live / SLC Live'. Follow up ~7 days.",
        },
        {
            1:  "SLC Live (The Gateway Olympic Legacy Plaza)",
            2:  "UT",
            3:  "Utah",
            4:  1,
            5:  "Yes",
            8:  "Dustin",
            9:  "dustin@parkcitylive.us",
            11: "Cold",
            12: "Email Sent",
            13: "2026-05-13",
            14: "2026-05-13",
            15: "2026-05-20",
            18: "Festival/event organizer at The Gateway Olympic Legacy Plaza, SLC. Pitched as Sierra-Wasatch snowcat moment for urban venue. Same contact manages Tahoe Live. Subject: 'Disco Carrot @ Tahoe Live / SLC Live'. Follow up ~7 days.",
        },
    ]

    for i, row_data in enumerate(new_rows):
        target_row = last_row + 1 + i
        for col, val in row_data.items():
            sheet.cell(row=target_row, column=col).value = val
        print(f"Added row {target_row}: {row_data[1]}")

    # Save updated xlsx
    out = io.BytesIO()
    wb.save(out)
    updated_bytes = out.getvalue()

    # Upload — overwrite existing file
    print("Uploading updated tracker...")
    media = MediaIoBaseUpload(io.BytesIO(updated_bytes), mimetype=TRACKER_MIME)
    drive.files().update(fileId=TRACKER_FILE_ID, media_body=media).execute()
    print("Done — tracker updated successfully.")


if __name__ == "__main__":
    main()
