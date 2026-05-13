"""
Weekly Gmail → TourTracker sync for Disco Carrot.

Reads the past 7 days of email, uses Claude to extract outreach activity,
then updates DiscoCarrot TourTracker v2.xlsx in Google Drive.

Required environment variables:
  ANTHROPIC_API_KEY
  GOOGLE_CLIENT_ID
  GOOGLE_CLIENT_SECRET
  GOOGLE_REFRESH_TOKEN
  GDRIVE_TRACKER_FILE_ID   (Drive file ID for DiscoCarrot TourTracker v2.xlsx)
  GDRIVE_ARCHIVE_FOLDER_ID (Drive folder ID where archives go)
"""

import io
import json
import os
import sys
import tempfile
from datetime import datetime, timedelta, timezone

import anthropic
import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload


# ── Constants ────────────────────────────────────────────────────────────────

TRACKER_FILE_ID = os.environ["GDRIVE_TRACKER_FILE_ID"]
ARCHIVE_FOLDER_ID = os.environ["GDRIVE_ARCHIVE_FOLDER_ID"]
TRACKER_FILENAME = "DiscoCarrot TourTracker v2.xlsx"
TRACKER_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Column indices in the Outreach Tracker sheet (0-based after header row)
COL = {
    "resort":         0,
    "state":          1,
    "region":         2,
    "tier":           3,
    "books_djs":      4,
    "season_open":    5,
    "season_close":   6,
    "contact_name":   7,
    "contact_email":  8,
    "contact_phone":  9,
    "outreach_type":  10,
    "status":         11,
    "date_first":     12,
    "date_last":      13,
    "next_followup":  14,
    "tour_window":    15,
    "booked_dates":   16,
    "notes":          17,
}

VALID_STATUSES = [
    "Not Contacted",
    "Email Sent",
    "Follow-up Sent",
    "Replied",
    "In Discussion",
    "Booked",
    "Declined",
    "On Hold",
]


# ── Google auth ───────────────────────────────────────────────────────────────

def _google_creds():
    creds = Credentials(
        token=None,
        refresh_token=os.environ["GOOGLE_REFRESH_TOKEN"],
        client_id=os.environ["GOOGLE_CLIENT_ID"],
        client_secret=os.environ["GOOGLE_CLIENT_SECRET"],
        token_uri="https://oauth2.googleapis.com/token",
        scopes=[
            "https://www.googleapis.com/auth/gmail.readonly",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    creds.refresh(Request())
    return creds


# ── Gmail helpers ─────────────────────────────────────────────────────────────

def fetch_week_emails(gmail):
    """Return plain-text summaries of all emails from the past 7 days."""
    since = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y/%m/%d")
    query = f"after:{since}"

    result = gmail.users().messages().list(userId="me", q=query, maxResults=200).execute()
    messages = result.get("messages", [])
    print(f"Found {len(messages)} emails in the past 7 days.")

    summaries = []
    for msg_ref in messages:
        msg = gmail.users().messages().get(
            userId="me", id=msg_ref["id"], format="metadata",
            metadataHeaders=["From", "To", "Subject", "Date"],
        ).execute()

        headers = {h["name"]: h["value"] for h in msg["payload"]["headers"]}
        snippet = msg.get("snippet", "")
        summaries.append(
            f"From: {headers.get('From','')}\n"
            f"To: {headers.get('To','')}\n"
            f"Subject: {headers.get('Subject','')}\n"
            f"Date: {headers.get('Date','')}\n"
            f"Snippet: {snippet}\n"
        )

    return summaries


# ── Claude analysis ───────────────────────────────────────────────────────────

def analyze_emails_with_claude(email_summaries: list[str]) -> list[dict]:
    """
    Ask Claude to extract outreach activity from emails and return structured
    tracker updates.

    Returns a list of update dicts, each with keys matching COL names plus
    only the fields that should change.
    """
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

    email_block = "\n---\n".join(email_summaries) if email_summaries else "(no emails)"

    prompt = f"""You are helping update the Disco Carrot tour booking tracker.

Disco Carrot is a DJ/live music act that books performances at ski resorts.
The tracker spreadsheet has one row per resort and tracks outreach status.

Valid status values:
{json.dumps(VALID_STATUSES, indent=2)}

Here are all emails from the past 7 days:

{email_block}

Analyze these emails and identify any outreach activity related to ski resort bookings.
For each resort where something happened, return a JSON array of update objects.

Each update object must have:
- "resort" (string): exact resort name as it appears in the tracker
- "status" (string): one of the valid statuses above, only if status changed
- "contact_name" (string, optional): if a contact was identified
- "contact_email" (string, optional): if a contact email was found
- "contact_phone" (string, optional): if a phone number was found
- "date_last" (string, optional): date of most recent contact in YYYY-MM-DD format
- "date_first" (string, optional): date of first contact if this is the first outreach, YYYY-MM-DD
- "next_followup" (string, optional): suggested follow-up date in YYYY-MM-DD format
- "outreach_type" (string, optional): "Cold", "Warm", "Referral", etc.
- "booked_dates" (string, optional): if dates were confirmed
- "notes" (string, optional): any relevant notes to append

Return ONLY a valid JSON array. If no relevant outreach activity was found, return [].
Do not include any explanation outside the JSON."""

    response = client.messages.create(
        model="claude-opus-4-7",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = response.content[0].text.strip()
    # Strip markdown fences if present
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1].rsplit("```", 1)[0]

    updates = json.loads(raw)
    print(f"Claude identified {len(updates)} tracker update(s).")
    return updates


# ── XLSX helpers ──────────────────────────────────────────────────────────────

def download_tracker(drive) -> bytes:
    request = drive.files().get_media(fileId=TRACKER_FILE_ID)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


def apply_updates(xlsx_bytes: bytes, updates: list[dict]) -> bytes:
    """Apply Claude's updates to the xlsx and return the modified bytes."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    # Find the Outreach Tracker sheet
    sheet = None
    for name in wb.sheetnames:
        if "outreach" in name.lower() or "tracker" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    # Build a lookup: resort name (lower) → row index (1-based, skipping header)
    header_row = 1
    resort_to_row = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=False):
        cell_val = row[COL["resort"]].value
        if cell_val:
            resort_to_row[str(cell_val).strip().lower()] = row

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for update in updates:
        resort_key = update.get("resort", "").strip().lower()
        row = resort_to_row.get(resort_key)
        if row is None:
            print(f"  ⚠ No tracker row found for resort: {update.get('resort')!r}")
            continue

        def set_cell(col_key, value):
            if value is not None:
                row[COL[col_key]].value = value

        set_cell("status", update.get("status"))
        set_cell("contact_name", update.get("contact_name"))
        set_cell("contact_email", update.get("contact_email"))
        set_cell("contact_phone", update.get("contact_phone"))
        set_cell("date_last", update.get("date_last") or today)
        set_cell("next_followup", update.get("next_followup"))
        set_cell("outreach_type", update.get("outreach_type"))
        set_cell("booked_dates", update.get("booked_dates"))

        # Append notes rather than overwrite
        if update.get("notes"):
            existing = row[COL["notes"]].value or ""
            separator = "\n" if existing else ""
            row[COL["notes"]].value = f"{existing}{separator}[{today}] {update['notes']}"

        # Set date_first only if it's currently empty
        if update.get("date_first") and not row[COL["date_first"]].value:
            row[COL["date_first"]].value = update["date_first"]

        print(f"  ✓ Updated: {update.get('resort')} → {update.get('status', '(no status change)')}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Drive upload helpers ──────────────────────────────────────────────────────

def overwrite_tracker(drive, xlsx_bytes: bytes) -> bool:
    """Try to update the existing tracker file. Returns True on success."""
    try:
        media = MediaIoBaseUpload(io.BytesIO(xlsx_bytes), mimetype=TRACKER_MIME)
        drive.files().update(fileId=TRACKER_FILE_ID, media_body=media).execute()
        print("Tracker overwritten successfully.")
        return True
    except HttpError as e:
        print(f"Could not overwrite tracker (HTTP {e.status_code}): {e}")
        return False


def archive_and_upload_new(drive, xlsx_bytes: bytes):
    """Move current tracker to archive folder, then upload fresh copy."""
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    archive_name = f"DiscoCarrot TourTracker v2 — archive {timestamp}.xlsx"

    # Copy original to archive folder with timestamped name
    drive.files().copy(
        fileId=TRACKER_FILE_ID,
        body={"name": archive_name, "parents": [ARCHIVE_FOLDER_ID]},
    ).execute()
    print(f"Archived original as: {archive_name}")

    # Upload new file to same parent folder as the original
    original_meta = drive.files().get(
        fileId=TRACKER_FILE_ID, fields="parents"
    ).execute()
    parents = original_meta.get("parents", [])

    media = MediaIoBaseUpload(io.BytesIO(xlsx_bytes), mimetype=TRACKER_MIME)
    new_file = drive.files().create(
        body={"name": TRACKER_FILENAME, "parents": parents},
        media_body=media,
        fields="id, name",
    ).execute()
    print(f"Uploaded new tracker: {new_file['name']} (id={new_file['id']})")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"=== Disco Carrot TourTracker sync — {datetime.now(timezone.utc).isoformat()} ===")

    creds = _google_creds()
    gmail = build("gmail", "v1", credentials=creds)
    drive = build("drive", "v3", credentials=creds)

    # 1. Read emails
    email_summaries = fetch_week_emails(gmail)

    # 2. Analyze with Claude
    updates = analyze_emails_with_claude(email_summaries)

    if not updates:
        print("No outreach activity detected. Tracker unchanged.")
        sys.exit(0)

    # 3. Download tracker
    print("Downloading tracker from Drive...")
    xlsx_bytes = download_tracker(drive)

    # 4. Apply updates
    updated_bytes = apply_updates(xlsx_bytes, updates)

    # 5. Upload — try overwrite first, fall back to archive + new
    if not overwrite_tracker(drive, updated_bytes):
        archive_and_upload_new(drive, updated_bytes)

    print("=== Sync complete ===")


if __name__ == "__main__":
    main()
